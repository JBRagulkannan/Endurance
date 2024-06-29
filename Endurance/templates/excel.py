import openpyxl
from openpyxl.cell import cell

# wb = openpyxl.Workbook()
#
# sheet1 = wb.active
#
# sheet1['A1'] = 87
# sheet1['A2'] = "New DATA"
# wb.save('example_excel.xlsx')
# <--------------------->
wb = openpyxl.load_workbook('example_excel.xlsx')
sheet = wb.active

data = (
    (11,12,13),
    (11,12,13),
    (11,12,13),
    (11,12,13),
    (11,12,13),("ragul",23,22)
)

for i in data:
    sheet.append(i)

a = sheet['E7']
print(a.value)
# cell = sheet.cell(row=7,colume=7)
# cell.font = cell.font.copy(bold=True)
sheet['E8'] = "Kannan"
wb.save('example_excel.xlsx')