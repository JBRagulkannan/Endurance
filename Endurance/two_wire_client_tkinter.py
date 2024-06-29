import binascii
from tkinter import *
import tkinter as tk
from tkinter import ttk
import ttkbootstrap as tb
from pip._internal.cli.cmdoptions import progress_bar
from ttkbootstrap.constants import *
from ttkbootstrap import Style
import os.path
from PIL import Image, ImageTk
import tkinter.messagebox as messagebox
import serial
import serial.tools.list_ports
import mysql.connector
import datetime
import time
import xlsxwriter
import openpyxl
from openpyxl import Workbook

fp_dict = {"FP16":"0","FP1":"1","FP2":"2","FP3":"3","FP4":"4","FP5":"5","FP6":"6","FP7":"7","FP8":"8",
               "FP9":"9","FP10":"a","FP11":"b","FP12":"c","FP13":"d","FP14":"e","FP15":"f"}

eot_fp_dict = {"FP16":"1","FP1":"2","FP2":"3","FP3":"4","FP4":"5","FP5":"6","FP6":"7","FP7":"8","FP8":"9",
               "FP9":"a","FP10":"b","FP11":"c","FP12":"d","FP13":"e","FP14":"f","FP15":"0"}

val_dict = {"0": "e0", "1": "e1", "2": "e2", "3": "e3", "4": "e4", "5": "e5", "6": "e6", "7": "e7", "8": "e8","9": "e9"}

grade_dict = {"g1": "e0", "g2": "e1", "g3": "e2", "g4": "e3", "g5": "e4", "g6": "e5", "g7": "e6", "g8": "e7","g9": "e8",
              "g10": "e9", "g11": "ea", "g12": "eb", "g13": "ec", "g14": "ed", "g15": "ee", "g16": "ef"}

saved_data = {}

def save_final_val_excel(final_dateTime, final_vol, final_sale, final_ppu, final_billno, final_err, final_crc_err):
    wb = openpyxl.load_workbook('TOT_Save.xlsx')
    sheet = wb.active
    last_row = sheet.max_row
    print(last_row)
    init_dateTime = sheet["A" + str(last_row)].value
    init_ppu = sheet["C" + str(last_row)].value
    init_vol = sheet["E" + str(last_row)].value
    init_sale = sheet["H" + str(last_row)].value
    init_bill = sheet["K" + str(last_row)].value
    init_err = sheet["M" + str(last_row)].value
    init_cr_err = sheet["O" + str(last_row)].value
    sheet["B" + str(last_row)] = final_dateTime
    sheet["H" + str(last_row)] = final_ppu
    sheet["J" + str(last_row)] = final_vol
    sheet["M" + str(last_row)] = final_sale
    sheet["P" + str(last_row)] = final_billno
    sheet["R" + str(last_row)] = final_err
    sheet["T" + str(last_row)] = final_crc_err
    wb.save('TOT_Save.xlsx')


def save_excel(sheet="",intDt=0,finDt=0,intvol=0,intsale=0,intppu=0,intbillno=0,interr=0,
               intcrcerr=0,finvol=0,finsale=0,finppu=0,finbillno=0,finerr=0,fincrcerr=0,power_on=0,power_off=0,
               nozzle_on=0,nozzle_off=0):
    # wb = Workbook()
    wb = openpyxl.load_workbook('TOT_Save.xlsx')
    sheet = wb[sheet]

    # int_vol = 123.09, 848, 98497
    sheet['A2'] = "Initial Date"
    sheet['B2'] = "Final Date"
    sheet['C2'] = "Power ON"
    sheet['D2'] = "Power OFF"
    sheet['E2'] = "Nozzle ON"
    sheet['F2'] = "Nozzle OFF"
    sheet['G2'] = "Initial PPU"
    sheet['H2'] = "Final PPU"
    sheet['I2'] = "Initial Vol"
    sheet['J2'] = "Final Vol"
    sheet['K2'] = "Vol diff"
    sheet['L2'] = "Initial sale"
    sheet['M2'] = "Final sale"
    sheet['N2'] = "Sale diff"
    sheet['O2'] = "Initial Bill NO"
    sheet['P2'] = "Final Bill NO"
    sheet['Q2'] = "Initial ERR"
    sheet['R2'] = "Final ERR"
    sheet['S2'] = "Initial CRT_ERR"
    sheet['T2'] = "Final CRT_ERR"

    if saved_data:
        if saved_data["power_fail"] == "YES":
            power_on = str(saved_data["power_time1"]) + str(saved_data["power_unit1"])
            power_off = str(saved_data["power_time2"]) + str(saved_data["power_unit2"])
        if saved_data["long"] != "YES":
            nozzle_on = "Long Endurance"
            nozzle_off = "Long Endurance"
        else:
            nozzle_on = str(saved_data["nozzle_on_time"]) + str(saved_data["nozzle_on_unit"])
            nozzle_off = str(saved_data["nozzle_off_time"]) + str(saved_data["nozzle_off_unit"])

    sheet.append({'A': intDt,'B': finDt, 'C': power_on, 'D': power_off,'E': nozzle_on,'F': nozzle_off,
                  'G': intppu,'H': finppu,'I': intvol, 'J': finvol,'K': (finvol - intvol),'L': intsale,
                  'M': finsale, 'N': (finsale - intsale),  'O': intbillno,'P': finbillno, 'Q': interr,
                  'R': finerr,'S': intcrcerr,'T': fincrcerr})

    wb.save('TOT_Save.xlsx')




def open_popup():
    top = Toplevel(root)
    top.geometry("500x320")  # Smaller size
    top.title("Power and Nozzle ON/OFF Timing")

    # Power Section
    power_frame = LabelFrame(top, text="Power Fail", padx=5, pady=5)  # Reduced padding
    power_frame.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")  # Reduced padding

    Checkbutton1 = IntVar()
    Checkbutton2 = IntVar()
    long_endurance_var = IntVar()

    def toggle_checkbutton1():
        if Checkbutton1.get() == 1:
            Checkbutton2.set(0)
            toggle_entries(textbox1, textbox2, 1)
        else:
            toggle_entries(textbox1, textbox2, 0)
        validate_inputs()

    def toggle_checkbutton2():
        if Checkbutton2.get() == 1:
            Checkbutton1.set(0)
            toggle_entries(textbox1, textbox2, 0)  # Ensure this disables the text boxes
        else:
            toggle_entries(textbox1, textbox2, 0)  # Ensure this disables the text boxes
        validate_inputs()

    def toggle_endurance():
        print(long_endurance_var.get())
        if long_endurance_var.get() == 1:
            toggle_entries(textbox3, textbox4, 0)  # Ensure this disables the text boxes
        else:
            toggle_entries(textbox3, textbox4, 1)  # Ensure this disables the text boxes
        validate_inputs()

    Checkbutton1_widget = Checkbutton(power_frame, text="YES",
                                      variable=Checkbutton1,
                                      onvalue=1,
                                      offvalue=0,
                                      height=1,  # Reduced height
                                      width=5,  # Reduced width
                                      command=toggle_checkbutton1)
    Checkbutton2_widget = Checkbutton(power_frame, text="NO",
                                      variable=Checkbutton2,
                                      onvalue=1,
                                      offvalue=0,
                                      height=1,  # Reduced height
                                      width=5,  # Reduced width
                                      command=toggle_checkbutton2)

    Checkbutton1_widget.grid(row=0, column=1, padx=5, pady=5)
    Checkbutton2_widget.grid(row=0, column=2, padx=5, pady=5)

    def validate_number(P):
        return P.isdigit() or P == ""

    vcmd = (top.register(validate_number), '%P')

    textbox1 = Entry(power_frame, width=10, state=tk.DISABLED, validate='key', validatecommand=vcmd)  # Adjusted width
    textbox1.grid(row=1, column=1, padx=5, pady=5)

    dropdown1_var = StringVar(top)
    dropdown1_var.set("Sec")
    dropdown1 = OptionMenu(power_frame, dropdown1_var, "MileSec", "Sec", "Min", "Hr")
    dropdown1.grid(row=1, column=2, padx=5, pady=5)

    textbox2 = Entry(power_frame, width=10, state=tk.DISABLED, validate='key', validatecommand=vcmd)  # Adjusted width
    textbox2.grid(row=1, column=3, padx=5, pady=5)

    dropdown2_var = StringVar(top)
    dropdown2_var.set("Sec")
    dropdown2 = OptionMenu(power_frame, dropdown2_var, "MileSec", "Sec", "Min", "Hr")
    dropdown2.grid(row=1, column=4, padx=5, pady=5)

    # Nozzle Section
    nozzle_frame = LabelFrame(top, text="Nozzle", padx=5, pady=5)  # Reduced padding
    nozzle_frame.grid(row=1, column=0, padx=5, pady=5, sticky="nsew")  # Reduced padding

    nozzle_label_on = Label(nozzle_frame, text="ON", padx=5, pady=5)  # Reduced padding
    nozzle_label_on.grid(row=0, column=0)

    dropdown1_var_on = StringVar(top)
    dropdown1_var_on.set("Sec")
    dropdown1_on = OptionMenu(nozzle_frame, dropdown1_var_on, "MileSec", "Sec", "Min", "Hr")
    dropdown1_on.grid(row=0, column=2, padx=5, pady=5)

    textbox3 = Entry(nozzle_frame, width=10, validate='key', validatecommand=vcmd)  # Adjusted width
    textbox3.grid(row=0, column=1, padx=5, pady=5)

    nozzle_label_off = Label(nozzle_frame, text="OFF", padx=5, pady=5)  # Reduced padding
    nozzle_label_off.grid(row=0, column=3)

    textbox4 = Entry(nozzle_frame, width=10, validate='key', validatecommand=vcmd)  # Adjusted width
    textbox4.grid(row=0, column=4, padx=5, pady=5)

    dropdown1_var_off = StringVar(top)
    dropdown1_var_off.set("Sec")
    dropdown1_off = OptionMenu(nozzle_frame, dropdown1_var_off, "MileSec", "Sec", "Min", "Hr")
    dropdown1_off.grid(row=0, column=5, padx=5, pady=5)

    long_endurance_label = Label(nozzle_frame, text="Long Endurance", padx=5, pady=5)
    long_endurance_label.grid(row=1, column=2, sticky='w')

    long_endurance_checkbox = Checkbutton(nozzle_frame,
                                          variable=long_endurance_var,
                                          onvalue=1,
                                          offvalue=0,
                                          height=1,  # Reduced height
                                          width=5,  # Reduced width
                                          command=toggle_endurance)
    long_endurance_checkbox.grid(row=1, column=3, padx=5, pady=5, sticky='w')

    # Function to toggle state of textbox1 and textbox2
    def toggle_entries(entry1, entry2, check_value):
        if check_value == 1:  # "YES" is checked
            entry1.config(state='normal')
            entry2.config(state='normal')
        else:  # "NO" is checked
            entry1.config(state='disabled')
            entry2.config(state='disabled')

    # long_endurance_checkbox.config(command=lambda: toggle_entries(textbox3, textbox4, not long_endurance_var.get()))

    # Function to validate if all required inputs are filled
    def validate_inputs():
        power_fail_selected = Checkbutton1.get() == 1 or Checkbutton2.get() == 1
        power_times_filled = (
                    textbox1.get().isdigit() and textbox2.get().isdigit()) if Checkbutton1.get() == 1 else True
        nozzle_times_filled = (textbox3.get().isdigit() and textbox4.get().isdigit())
        nozzle_times_filled1 = long_endurance_var.get() == 1
        print(nozzle_times_filled1)
        print(long_endurance_var.get())
        if power_fail_selected and power_times_filled and (nozzle_times_filled or nozzle_times_filled1):
            save_button.config(state='normal')
        else:
            save_button.config(state='disabled')

    # Bind validation function to Entry widgets
    for widget in [textbox1, textbox2, textbox3, textbox4]:
        widget.bind("<KeyRelease>", lambda event: validate_inputs())

    # Adjust row and column weights for grid resizing
    top.grid_rowconfigure(0, weight=1)
    top.grid_rowconfigure(1, weight=1)
    top.grid_columnconfigure(0, weight=1)

    def save_data():
        power_fail = "YES" if Checkbutton1.get() == 1 else "NO"
        power_time1 = textbox1.get()
        power_unit1 = dropdown1_var.get()
        power_time2 = textbox2.get()
        power_unit2 = dropdown2_var.get()

        nozzle_on_time = textbox3.get()
        nozzle_on_unit = dropdown1_var_on.get()
        nozzle_off_time = textbox4.get()
        nozzle_off_unit = dropdown1_var_off.get()
        long_endurance_var = "longEndurance"

        saved_data.update({"power_fail": power_fail,"power_time1": power_time1,"power_unit1": power_unit1,
                          "power_time2": power_time2, "power_unit2": power_unit2, "nozzle_on_time": nozzle_on_time,
                          "nozzle_on_unit": nozzle_on_unit, "nozzle_off_time": nozzle_off_time,
                          "nozzle_off_unit": nozzle_off_unit, "long": long_endurance_var})

        print(f"Power Fail: {power_fail}")
        if power_fail == "YES":
            print(f"Power ON Time: {power_time1} {power_unit1}")
            print(f"Power OFF Time: {power_time2} {power_unit2}")
        print(f"Nozzle ON Time: {nozzle_on_time} {nozzle_on_unit}")
        print(f"Nozzle OFF Time: {nozzle_off_time} {nozzle_off_unit}")
        print(f"Nozzle Time: {long_endurance_var}")

        top.destroy()

    save_button = Button(top, text="Save", width=15, state='disabled', command=save_data)  # Adjusted width
    save_button.grid(row=2, column=0, padx=5, pady=10, columnspan=2)


# image = Image.open('C:\\Users\\Ragul.Kannan\\Downloads\\1705941005496.jpg')
# image = image.resize((320, 280))
# photo = ImageTk.PhotoImage(image)
#
# background_label = Label(root, image=photo)
# background_label.place(x=0, y=0, relwidth=1, relheight=1)
con = mysql.connector.Connect(host="localhost", user="root", password="12345678", database="Endurance_db")
print(con)

def upsert():
    sqlCursor = con.cursor()
    project_name = "hpcl"
    nozzlename = "A1"
    int_vol = "123"
    int_vol1 = "456"
    int_sale = "1223"
    int_sale1 = "14789"
    int_ppu = "12.34"
    int_ppu1 = "24.68"
    int_billno = "111"
    int_billno1 = "134"
    int_err = "E09"
    int_err1 = "E09"
    int_crc_err = "E139"
    int_crc_err1 = "E139"

    values = [
        # ('Project A', 'Config A', 'Nozzle 1', 1.23, 100, 123.0, 'B001', 0, 0),
        ('Project B', 'Config B', 'Nozzle 2', 12.34, 150, 145.0, 'B002', 0, 0),
        # Add more tuples/lists as needed
    ]


    upsert_query = "INSERT INTO Endurance (project_name, config, nozzlename, initial_PPU, initial_volume, initial_sale, initial_Billno, initial_Err, initial_critical_Err)VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s) ON DUPLICATE KEY UPDATE initial_PPU=VALUES(initial_PPU)"
    # upsert_query = "INSERT INTO Endurance (project_name, config, nozzlename, initial_PPU, initial_volume, initial_sale, initial_Billno, initial_Err, initial_critical_Err) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)"
    val = tuple(values)
    sqlCursor.executemany(upsert_query, values)
    con.commit()
    sqlCursor.close()
    print(sqlCursor)

# project_name = "hpcl"
# nozzlename = "A1"
# int_vol = 123
# int_vol1 = 456
# int_sale = 1223
# int_sale1 = 14789
# int_ppu = 12.34
# int_ppu1 = 24.68
# int_billno = 111
# int_billno1 = 134
# int_err = "E09"
# int_err1 = "E09"
# int_crc_err = "E139"
# int_crc_err1 = "E139"



def insert(initial,final,curr_date,isInitial):
    print("workinf strat")
    cursor = con.cursor()
    data = {
        'initial':initial,
        'final':final,
        'curr_date':curr_date
    }
    if isInitial:
        query = "insert into users(initial,final,curr_date)VALUES (%(initial)s, %(final)s, %(curr_date)s) ON DUPLICATE KEY UPDATE initial=Values(initial),curr_date = Values(curr_date)"
    else:
        query = "insert into users(initial,final,curr_date)VALUES (%(initial)s, %(final)s, %(curr_date)s) ON DUPLICATE KEY UPDATE final=Values(final),curr_date = Values(curr_date)"
    cursor.execute(query,data)
    con.commit()
    cursor.close()
    print("working fin")

def get_usb_ports():
    usb_ports = []
    ports = serial.tools.list_ports.comports()
    # print(ports)
    for port in ports:
        if 'USB' in port.description:
            usb_ports.append(port.device)
            print(usb_ports)
    return usb_ports

def ser_initialize():
    global ser
    print(clicked.get())
    ser = serial.Serial(
        port=clicked.get(),
        baudrate=5787,
        parity=serial.PARITY_EVEN,
        stopbits=serial.STOPBITS_ONE,
        bytesize=serial.EIGHTBITS,
        timeout=3)
    ser.close()

def poll(data):
    ser.open()
    print(data)
    fp_data = "0" + fp_dict.get(data)
    ser.write(binascii.unhexlify(fp_data))
    print(type(fp_data))
    fp_id = fp_data[1]
    var_out = binascii.hexlify(ser.read())
    fp_id_out = str(var_out).replace("b'","")
    fp_id_out = fp_id_out.replace("'","")
    ser.close()
    if fp_id_out:
        print("fp_state: " + fp_id_out)
        print(data + " is polling")
        return True
    else:
        print(data + " is not polling")
        return False



def get_totals(data, status_std = "std"):
    ser.open()
    fp_data = "5" + fp_dict.get(data)
    ser.write(binascii.unhexlify(fp_data))
    print(type(fp_data))
    #fp_id = fp_data[1]
    var_out = binascii.hexlify(ser.readline())

    fp_id_out = str(var_out).replace("b'", "")
    fp_id_out = fp_id_out.replace("'", "")
    ser.close()
    if len(fp_id_out) > 2:
        if status_std.lower() == "std":
            str1 = fp_id_out[8:24]
            str_sale = fp_id_out[34:58]
            str_ppu = fp_id_out[60:72]
            # str_ERR = fp_id_out[24:27]
            # str_CR_ERR = fp_id_out[24:27]



            print(str1,"volume")
            print(str_sale,"sale")
            print(str_ppu,"PPU dummy")
            # print(str_ERR,"ERR")
            # print(str_CR_ERR,"CR_ERR")



        elif status_std.lower() == "ext":
            str1 = fp_id_out[8:32]
            print(str1)
        else:
            print("Wrong Value mentioned in status_std")
            return False
        reversed_str = str1[::-1]
        removed_e = reversed_str.replace("e", "")
        print("Value is:", removed_e)
        final_output = float(removed_e[:-2] + "." + removed_e[-2:])
        print(final_output)

        reversed_str = str_sale[::-1]
        removed_e = reversed_str.replace("e", "")
        print("Value is:", removed_e)
        final_output_sale = float(removed_e[:-2] + "." + removed_e[-2:])
        print(final_output_sale)

        reversed_str = str_ppu[::-1]
        removed_e = reversed_str.replace("e", "")
        print("Value is:", removed_e)
        final_output_PPU = float(removed_e[:-2] + "." + removed_e[-2:])
        print(final_output_PPU,"ppu")

        # reversed_str = str_ERR[::-1]
        # removed_e = reversed_str.replace("e", "")
        # print("Value is:", removed_e)
        # final_output_Err = float(removed_e[:-2] + "." + removed_e[-2:])
        # print(final_output_Err, "ERR")
        #
        # reversed_str = str_ERR[::-1]
        # removed_e = reversed_str.replace("e", "")
        # print("Value is:", removed_e)
        # final_output_CR_ERR = float(removed_e[:-2] + "." + removed_e[-2:])
        # print(final_output_CR_ERR, "CR_ERR")

        return final_output, final_output_sale, final_output_PPU

    else:
        print("fp_state: ", fp_id_out)
        print("Length of Response is not expected. Check the value")
        return False


def get_ERR(data, err_type):
    err_cmd = ""
    if err_type == "critical_err":
        err_cmd = "ffe3fee0e5e3e0e0e0e0e0e0fbedf0"
    elif err_type == "partial_err":
        err_cmd = "ffe3fee0e5e3e0e3e0e0e0e0fbeaf0"

    ser.open()
    fp_data = "2" + fp_dict.get(data)
    ser.write(binascii.unhexlify(fp_data))
    print(type(fp_data))
    print(fp_data)
    # fp_id = fp_data[1]
    var_out = binascii.hexlify(ser.readline())
    print(var_out)
    fp_id_out = str(var_out).replace("b'", "")
    fp_id_out = fp_id_out.replace("'", "")
    if fp_id_out:
        ser.write(binascii.unhexlify(err_cmd))
        var_out = binascii.hexlify(ser.readline())
        print(var_out)

        fp_id_out = str(var_out).replace("b'", "")
        fp_id_out = fp_id_out.replace("'", "")
        ser.close()
        if len(fp_id_out) > 2:
            if err_type.lower() == "partial_err":
                str_ERR = fp_id_out[46:54]
                print(str_ERR,"ERR")
                # reversed_str = str_ERR[::-1]
                final_output_Err = str_ERR.replace("b", "")
                print(final_output_Err, "ERR")
                return final_output_Err
            elif err_type.lower() == "critical_err":
                str_CR_ERR = fp_id_out[46:54]
                print(str_CR_ERR,"CR_ERR")
                # reversed_str = str_CR_ERR[::-1]
                final_output_CR_ERR = str_CR_ERR.replace("b", "")
                # print("Value is:", removed_e)
                # final_output_CR_ERR = float(removed_e[:-2] + "." + removed_e[-2:])
                print(final_output_CR_ERR, "CR_ERR")
                return final_output_CR_ERR
            else:
                print("Wrong Value mentioned in err_type")
                return False
    else:
        print("Serial response not received")
        ser.close()
        return False

def get_billno(data):
    billno_cmd = "ffe3fee0e5e3e1e0e0e0e0e0fbecf0"
    ser.open()
    fp_data = "2" + fp_dict.get(data)
    ser.write(binascii.unhexlify(fp_data))
    print(type(fp_data))
    print(fp_data)
    # fp_id = fp_data[1]
    var_out = binascii.hexlify(ser.readline())
    print(var_out)
    fp_id_out = str(var_out).replace("b'", "")
    fp_id_out = fp_id_out.replace("'", "")
    if fp_id_out:
        ser.write(binascii.unhexlify(billno_cmd))
        var_out = binascii.hexlify(ser.readline())
        print(var_out)

        fp_id_out = str(var_out).replace("b'", "")
        fp_id_out = fp_id_out.replace("'", "")
        ser.close()
        if len(fp_id_out) > 2:
            # if err_type.lower() == "partial_err":
                str_ERR = fp_id_out[48:60]
                print(str_ERR, "billno")
                # reversed_str = str_ERR[::-1]
                final_output_Err = str_ERR.replace("b", "")
                print("billno",final_output_Err)
                return final_output_Err

def start_progress_bar():
    progress_bar = ttk.Progressbar(root, mode='indeterminate')
    progress_bar.grid(row=7, column=5, pady=10, padx=10)
    progress_bar.start()
def take_initial_value():
    # start_progress_bar()
    # global int_dateTime, int_vol, int_sale, int_ppu, int_billno, int_err, int_crc_err
    # dummy_int_a1 = 109.984
    # dummy_int_a2 = 109.984
    # dummy_int_ppu = 109.984
    # dummy_int_err = "E09"
    # dummy_int_crc_err = "E01"
    # print(dummy_int_a1,dummy_int_a2,dummy_int_ppu,dummy_int_err,dummy_int_crc_err)
    # data_list = [{"DA": "13 / MAY / 12","int_vol":1223 ,"int_sale":300.25}]

    # a = 2
    # for dummy_excel in data_list:
    #     a = a + 1
    #     save_excel(dummy_excel["DA"], dummy_excel["int_vol"], dummy_excel["int_sale"], a)
    global saved_data
    upsert()
    print("i am working")
    ser_initialize()
    print("Taking initial value")
    # ret = poll("FP1")
    # ret1 = poll("FP2")
    ret = True  #dummy need to rechange
    ret1 = True
    i = 123.23
    print(i)
    insert(i, None, cur_datetime, True)
    print(i)

    # int_dateTime = str(datetime.datetime.now())
    # int_dateTime1 = int_dateTime.split(" ")[0]
    # print(int_dateTime)
    # if int_dateTime == int_dateTime1:
    #     messagebox.showerror("Alert", int_dateTime + "The same data can be found on that date in the Excel sheet. If yes, the same date will be modified.")
    #     root.destroy()
    # else:
    #     save_excel(intDt=int_dateTime)

    if ret == True and ret1 == True:
        # fp1_totals = get_totals("FP1")
        # fp2_totals = get_totals("FP2")
        # fp1_get_err = get_ERR("FP1",err_type="partial_err")
        # fp1_get_crc = get_ERR("FP1",err_type="critical_err")
        # fp2_get_err = get_ERR("FP2",err_type="partial_err")
        # fp2_get_crc = get_ERR("FP2",err_type="critical_err")
        # fp1_bill_no = get_billno("FP1")
        # fp2_bill_no = get_billno("FP2")
        fp1_totals = [12,1234,12.34]
        fp2_totals = [14,1577,24.68]
        fp1_get_err = "E09"
        fp1_get_crc = "E139"
        fp2_get_err = "E09"
        fp2_get_crc = "E09"
        fp1_bill_no = 187
        fp2_bill_no = 111

        int_dateTime = datetime.datetime.now()
        int_vol = fp1_totals[0]
        int_vol1 = fp2_totals[0]
        int_sale = fp1_totals[1]
        int_sale1 = fp2_totals[1]
        int_ppu = fp1_totals[2]
        int_ppu1 = fp2_totals[2]
        int_billno = fp1_bill_no
        int_billno1 = fp2_bill_no
        int_err = fp1_get_err
        int_err1 = fp2_get_err
        int_crc_err = fp1_get_crc
        int_crc_err1 = fp2_get_crc


        # excel = save_excel()
        print("FP1 Totals: ", fp1_totals)
        print("FP1 partial_err: ", fp1_get_err)
        print("FP1 critical_err: ", fp1_get_crc)
        print("FP2 Totals: ", fp2_totals)
        print("FP2 partial_err: ", fp2_get_err)
        print("FP2 critical_err: ", fp2_get_crc)
        print("FP1 Bill No: ", fp1_bill_no)
        print("FP2 Bill No: ", fp2_bill_no)
        print("Saved Data:")

        def powerandnozzledata():
            for key, value in saved_data.items():
                print(f"{key}: {value}")

        save_excel(sheet="A1",intDt=int_dateTime, intvol=int_vol, intsale=int_sale, intppu=int_ppu,
                   intbillno=int_billno,interr=int_err,intcrcerr=int_crc_err)
        save_excel(sheet="A2",intDt=int_dateTime, intvol=int_vol1, intsale=int_sale1, intppu=int_ppu1,
                   intbillno=int_billno1, interr=int_err1, intcrcerr=int_crc_err1)

        # root.after(5000, lambda: stop_progress_bar(progress_bar))
        messagebox.showinfo("Success", "Initial values taken successfully across all stores.")


def stop_progress_bar(progress_bar):
            progress_bar.stop()
            progress_bar.grid_forget()

        # save_excel(
        #     intDt=saved_data.get("int_dateTime", int_dateTime),
        #     intvol=saved_data.get("int_vol", int_vol),
        #     intsale=saved_data.get("int_sale", int_sale),
        #     intppu=saved_data.get("int_ppu", int_ppu),
        #     intbillno=saved_data.get("int_billno", int_billno),
        #     interr=saved_data.get("int_err", int_err),
        #     intcrcerr=saved_data.get("int_crc_err", int_crc_err)
        # )



def take_final_value():
    print("Taking final value")
    ser_initialize()
    ret = poll("FP1")
    ret1 = poll("FP2")
    y = 999.98
    # save_excel(finvol=y,column=6) # <--dummy propose-->
    if ret == True and ret1 == True:
        fp1_totals = get_totals("FP1")
        fp2_totals = get_totals("FP2")
        fp1_get_err = get_ERR("FP1", err_type="partial_err")
        fp1_get_crc = get_ERR("FP1", err_type="critical_err")
        fp2_get_err = get_ERR("FP2", err_type="partial_err")
        fp2_get_crc = get_ERR("FP2", err_type="critical_err")
        fp1_bill_no = get_billno("FP1")
        fp2_bill_no = get_billno("FP2")

        final_dateTime = datetime.datetime.now()
        final_vol = fp1_totals[0]
        final_sale = fp1_totals[1]
        final_ppu = fp1_totals[2]
        final_billno = fp1_bill_no
        final_err = fp1_get_err
        final_crc_err = fp1_get_crc

        # excel = save_excel()
        print("FP1 Totals: ", fp1_totals)
        print("FP1 partial_err: ", fp1_get_err)
        print("FP1 critical_err: ", fp1_get_crc)
        print("FP2 Totals: ", fp2_totals)
        print("FP2 partial_err: ", fp2_get_err)
        print("FP2 critical_err: ", fp2_get_crc)
        print("FP1 Bill No: ", fp1_bill_no)
        print("FP2 Bill No: ", fp2_bill_no)

        save_final_val_excel(final_dateTime, final_vol, final_sale, final_ppu, final_billno, final_err, final_crc_err)


def save_and_exit():
    # Code to save and exit goes here
    root.destroy()

# title_lbl = Label(root, text='Regional Dispenser Endurance Tool V1.0', width=45, height=1)
# title_lbl.grid(row=0, column=0, columnspan=50, sticky="e")
#
# input_lbl = Label(root, text='Enter Your Project:')
# input_lbl.grid(row=1, column=4)
#
# input_txt = Entry(root)
# input_txt.grid(row=1, column=5)
#
# small_btn = Button(root, text='Go')
# small_btn.grid(row=1, column=6)
#
# clicked = StringVar()
# clicked.set("Select")
#
# clicked1 = StringVar()
# clicked1.set("Config")
#
# clicked2 = StringVar()
# clicked2.set("Power Time")
#
# clicked3 = StringVar()
# clicked3.set("Nozzle Time")
#
# optn_lbl = Label(root, text='Select Com Port')
# optn_lbl.grid(row=4, column=4, )
#
# drop = OptionMenu(root, clicked, *com_ports)
# drop.grid(row=4, column=5)
#
# empty_lbl = Label(root, text='')
# empty_lbl.grid(row=5, column=0, columnspan=50)
#
# optn_lbl = Label(root, text='Select Config')
# optn_lbl.grid(row=5, column=4, )
#
# drop = OptionMenu(root,clicked1, *config)
# drop.grid(row=5, column=5)
#
# empty_lbl = Label(root, text='')
# empty_lbl.grid(row=5, column=0, columnspan=50)
#
# optn_lbl = Label(root, text='Timeing')
# optn_lbl.grid(row=6, column=4, )
#
# # small_btn = Button(root, text='Enter the value')
# # small_btn.grid(row=6, column=5)
#
# button = ttk.Button(root, text="Enter the value", command=open_popup)
# button.grid(row=6, column=5)
#
# empty_lbl = Label(root, text='')
# empty_lbl.grid(row=5, column=0, columnspan=50)
#
# init_lbl = Label(root, text='Click to take Initial Value')
# init_lbl.grid(row=8, column=4)
#
# init_btn = Button(root, text='Initial', bd='5', command=take_initial_value)
# init_btn.grid(row=8, column=5)
#
# empty_lbl = Label(root, text='')
# empty_lbl.grid(row=8, column=0, columnspan=50)
#
# fin_lbl = Label(root, text='Click to take Final Value')
# fin_lbl.grid(row=9, column=4)
#
# fin_btn = Button(root, text='Final', bd='5', command=take_final_value)
# fin_btn.grid(row=9, column=5)
#
# empty_lbl = Label(root, text='')
# empty_lbl.grid(row=10, column=0, columnspan=50)
#
# save_btn = Button(root, text='Save & Exit', bd='5', command=save_and_exit)
# save_btn.grid(row=11, column=5)

def go_button_clicked():
    project_name = input_txt.get()
    if project_name:
        filename = f"{project_name}.xlsx"
        if os.path.exists(filename):
            messagebox.showerror("Error", f"Excel sheet '{filename}' already exists!")
        else:
            try:
                # Create Excel sheet with project name
                wb = openpyxl.Workbook()
                wb.save(filename)
                messagebox.showinfo("Success", f"Excel sheet '{filename}' created successfully!")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to create Excel sheet: {e}")
    else:
        messagebox.showwarning("Warning", "Please enter a project name.")




ser = None

cur_datetime = datetime.datetime.now().strftime('%Y-%m-%d')

style = Style(theme='superhero')  # Choose from themes: 'superhero', 'cosmo', 'darkly', etc.
root = style.master
root.title("Regional Dispenser Endurance Tool V1.0")
root.geometry("600x450")
top=None

com_ports = get_usb_ports()

if not com_ports:
    messagebox.showerror("Error", "Please Connect 2 wire Cable \U0001F923!")
    root.destroy()

config= ["MONO", "1P MPD", "2P MPD", "3P MPD", "4P MPD", "MONO", "DUO", "QUAD"]


title_lbl = ttk.Label(root, text='Regional Dispenser Endurance Tool V1.0', font=("Helvetica", 20, "bold"), anchor="center")
title_lbl.grid(row=0, column=0, columnspan=50, pady=20)

# Project input
input_lbl = ttk.Label(root, text='Enter Your Project:', font=("Helvetica", 14))
input_lbl.grid(row=1, column=4, pady=10, padx=10, sticky="e")

input_txt = ttk.Entry(root, width=30, font=("Helvetica", 12))
input_txt.grid(row=1, column=5, pady=10, padx=10)

small_btn = ttk.Button(root, text='Go', style='primary.TButton',command=go_button_clicked)
small_btn.grid(row=1, column=6, pady=10, padx=10)

# Option Menus
clicked = tk.StringVar(value="Select")
clicked1 = tk.StringVar(value="Config")
clicked2 = tk.StringVar(value="Power Time")
clicked3 = tk.StringVar(value="Nozzle Time")

# com_ports = ["COM1", "COM2", "COM3"]  # Example list of COM ports
# config = ["Config1", "Config2", "Config3"]  # Example list of configurations

optn_lbl = ttk.Label(root, text='Select Com Port', font=("Helvetica", 14))
optn_lbl.grid(row=4, column=4, pady=10, padx=10, sticky="e")

drop = ttk.OptionMenu(root, clicked, *com_ports)
drop.grid(row=4, column=5, pady=10, padx=10)

empty_lbl = ttk.Label(root, text='')
empty_lbl.grid(row=5, column=0, columnspan=50)

optn_lbl = ttk.Label(root, text='Select Config', font=("Helvetica", 14))
optn_lbl.grid(row=5, column=4, pady=10, padx=10, sticky="e")

drop = ttk.OptionMenu(root, clicked1, *config)
drop.grid(row=5, column=5, pady=10, padx=10)

empty_lbl = ttk.Label(root, text='')
empty_lbl.grid(row=5, column=0, columnspan=50)

optn_lbl = ttk.Label(root, text='Timing', font=("Helvetica", 14))
optn_lbl.grid(row=6, column=4, pady=10, padx=10, sticky="e")

button = ttk.Button(root, text="Enter the value", style='primary.TButton', command=open_popup)
button.grid(row=6, column=5, pady=10, padx=10)

empty_lbl = ttk.Label(root, text='')
empty_lbl.grid(row=5, column=0, columnspan=50)

init_lbl = ttk.Label(root, text='Click to take Initial Value', font=("Helvetica", 14))
init_lbl.grid(row=8, column=4, pady=10, padx=10, sticky="e")

init_btn = ttk.Button(root, text='Initial', style='success.TButton', command=take_initial_value)
init_btn.grid(row=8, column=5, pady=10, padx=10)

empty_lbl = ttk.Label(root, text='')
empty_lbl.grid(row=8, column=0, columnspan=50)

fin_lbl = ttk.Label(root, text='Click to take Final Value', font=("Helvetica", 14))
fin_lbl.grid(row=9, column=4, pady=10, padx=10, sticky="e")

fin_btn = ttk.Button(root, text='Final', style='warning.TButton', command=take_final_value)
fin_btn.grid(row=9, column=5, pady=10, padx=10)

empty_lbl = ttk.Label(root, text='')
empty_lbl.grid(row=10, column=0, columnspan=50)

save_btn = ttk.Button(root, text='Save & Exit', style='danger.TButton', command=save_and_exit)
save_btn.grid(row=11, column=5, pady=20, padx=10)

root.mainloop()
